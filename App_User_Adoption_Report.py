"""
App User Adoption Report

Reads user account exports, filters active users,
calculates app adoption rate by organisational unit,
and outputs a timestamped Excel report.
"""

import pandas as pd
import pandas.io.formats.excel
import os
import glob
import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

os.makedirs(OUTPUT_DIR, exist_ok=True)

COMPANY_FILTER = "ExampleCorp"
TEST_ACCOUNTS = ["Test User"]

pd.io.formats.excel.ExcelFormatter.header_style = None


def generate_report():
    all_files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))

    for file in all_files:
        df = pd.read_excel(file)

        # Remove test and internal accounts
        df = df[~df['Name'].isin(TEST_ACCOUNTS)]

        # Active users by organisational unit
        df_all = df[
            (df['Company'] == COMPANY_FILTER)
            & (df['IsActive'] == "Yes")
        ]
        df_all = df_all[['Name', 'EmployeeNumber', 'Phone',
                          'Email', 'Region', 'Area', 'Branch']]
        all_users = pd.pivot_table(
            df_all, index=['Region', 'Area', 'Branch'], aggfunc='count'
        )
        all_users = all_users.drop(
            ['Email', 'EmployeeNumber', 'Name'], axis=1
        )
        all_users = all_users.rename(columns={'Phone': 'Total Users'})

        # App users by organisational unit
        df_app = df[
            (df['Company'] == COMPANY_FILTER)
            & (df['AppUser'] == "Yes")
            & (df['IsActive'] == "Yes")
        ]
        df_app = df_app[['Name', 'EmployeeNumber', 'Phone',
                          'Email', 'Region', 'Area', 'Branch']]
        app_users = pd.pivot_table(
            df_app, index=['Region', 'Area', 'Branch'], aggfunc='count'
        )
        app_users = app_users.drop(
            ['Email', 'EmployeeNumber', 'Name'], axis=1
        )
        app_users = app_users.rename(columns={'Phone': 'App Users'})

        # Merge and calculate adoption rate
        merged = pd.merge(
            all_users, app_users,
            on=['Region', 'Area', 'Branch'], how='left'
        )
        merged['App Users'] = merged['App Users'].fillna(0).astype(int)
        merged['Total Users'] = merged['Total Users'].astype(int)
        merged['Adoption %'] = (
            (merged['App Users'] / merged['Total Users']) * 100
        ).round(0).fillna(0).astype(int).astype(str).add("%")

        # Full user list with app status (for detail sheet)
        df_detail = df[
            (df['Company'] == COMPANY_FILTER)
            & (df['IsActive'] == "Yes")
        ]
        df_detail = df_detail[[
            'Name', 'EmployeeNumber', 'Phone',
            'Email', 'Region', 'Area', 'Branch', 'AppUser'
        ]]

        # Write timestamped output
        now = datetime.datetime.now()
        timestamp = f"{now.year}-{now.month}-{now.day} {now.hour}-{now.minute}"
        output_path = os.path.join(
            OUTPUT_DIR, f"app_adoption_report  {timestamp}.xlsx"
        )
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_detail.to_excel(writer, sheet_name='user_list', index=False)
            merged.to_excel(writer, sheet_name='summary')
        
        # Style the output
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = load_workbook(output_path)
        
        # Style the summary sheet
        ws = wb['summary']
        
        # Header row — dark blue background, white text, bold
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows — borders + center alignment
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # Highlight adoption % column — light green for high, light red for low
        pct_col = None
        for cell in ws[1]:
            if 'pct' in str(cell.value).lower() or '%' in str(cell.value):
                pct_col = cell.column
                break
        
        if pct_col:
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                     min_col=pct_col, max_col=pct_col):
                for cell in row:
                    val = str(cell.value).replace('%', '').strip()
                    try:
                        if int(val) >= 75:
                            cell.fill = green_fill
                        elif int(val) <= 25:
                            cell.fill = red_fill
                    except ValueError:
                        pass
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 4
        
        # Style the user_list sheet too — just headers
        ws2 = wb['user_list']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        wb.save(output_path)

if __name__ == "__main__":
    generate_report()
